import openpyxl
from pathlib import Path
from sklearn.neural_network import MLPRegressor

main_file_prefix = "dane\\pozyxAPI_only_localization_measurement"
suffix = ".xlsx"
nr_of_records = 1540
train_x = []
train_x_ref = []
train_y = []
train_y_ref = []
test_x = []
test_y = []


def read_file(str_nr: str):
    xlsx_file = Path(main_file_prefix + str_nr + suffix)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    for row in sheet.iter_rows(2, nr_of_records + 1):
        train_x.append([row[4].value, row[5].value])
        # train_y.append(row[5].value)
        train_x_ref.append([row[6].value, row[7].value])
        # train_y_ref.append(row[7].value)

        # for i in range(4, 8):
        #     print(row[i].value, end=" ")


def read_file2():
    xlsx_file = Path('dane\\pozyxAPI_only_localization_dane_testowe_i_dystrybuanta' + suffix)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    for row in sheet.iter_rows(2, nr_of_records + 1):
        test_x.append([row[4].value, row[5].value])
        # test_y.append(row[5].value)
        # test_x.append(row[6].value)
        # test_y.append(row[7].value)
        # for i in range(4, 8):
        #     print(row[i].value, end=" ")


if __name__ == "__main__":
    print('hello world')
    for i in range(1, 13):
        read_file(str(i))
    read_file2()
    print(train_x[0], " ", train_x[1])
    # print(train_y[3])
    # print(test_x[3])
    # print(test_y[3])

    # mlp = MLPRegressor(hidden_layer_sizes=50, activation='tanh', solver='adam', alpha=0.0001, learning_rate_init=0.002,
    #                    max_iter=200, shuffle=True, tol=1, verbose=False, early_stopping=False)

    mlp = MLPRegressor(hidden_layer_sizes=(50,50,50), activation='relu', solver='adam', alpha=0.0001,
                       max_iter=200, shuffle=True, tol=1, verbose=False, early_stopping=False)

    # mlp.n_layers_ = 3
    mlp.out_activation_ = 'Zadanie2 - tanh, wrrrr'
    # mlp.n_outputs_ = 2
    print(mlp.out_activation_)
    print(mlp.__repr__())
    mlp.fit(train_x, train_x_ref)
    print(test_x)
    # print(train_x)
    # print(train_x_ref)
    print(mlp.predict(test_x))  # sprawdzenie
