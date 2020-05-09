from pathlib import Path

import openpyxl

prefix = 'dane\\pozyxAPI_only_localization'
main_file = prefix + "_measurement"
test_file = prefix + '_dane_testowe_i_dystrybuanta'
suffix = ".xlsx"
nr_of_records = 1540


class FileReader:

    def __init__(self):
        self.train_ref = []
        self.train = []
        self.test = []
        self.test_ref = []

    def read_train_files(self, nr: int):
        xlsx_file = Path(main_file + str(nr) + suffix)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        for row in sheet.iter_rows(2, nr_of_records + 1):
            self.train.append([row[4].value, row[5].value])
            self.train_ref.append([row[6].value, row[7].value])

    def read_test_file(self):
        xlsx_file = Path(test_file + suffix)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        for row in sheet.iter_rows(2, nr_of_records + 1):
            self.test.append([row[4].value, row[5].value])
            self.test_ref.append([row[4].value, row[5].value])

    def read_files(self):
        for i in range(1, 13):
            self.read_train_files(i)
        self.read_test_file()
        return self.get_all_data()

    def get_all_data(self):
        return self.train, self.train_ref, self.test, self.test_ref
